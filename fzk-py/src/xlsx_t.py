# -*- coding: utf-8 -*-  
from openpyxl import load_workbook
import json

course_names = ["物理","化学","生物","数学","历史","地理"]
school_type_names = ["小学","初中","高中"]
school_type_id_dict={"小学":21,"初中":31,"高中":34}
course_id_dict={"物理":16,"化学":17,"生物":18,"数学":14,"历史":21,"地理":20}

class Tx(object):

    def __init__(self, ygtx, k1ktx):
        self.ygtx = ygtx
        self.k1ktx = k1ktx

        
class TxTmp(object):
    def __init__(self, school_type_name, course_name, txs):
        self.school_type_name = school_type_name
        self.school_type_id = school_type_id_dict[school_type_name]
        self.course_name = course_name
        self.course_id = course_id_dict[course_name]
        self.txs = txs


class XslxVaue(object):

    def __init__(self, sheet_name:str, lines:[]):
        self.sheet_name = sheet_name
        self.lines = lines;

    @staticmethod
    def list_to_json(values):
        result = json.dumps(values, default=lambda o: o.__dict__, ensure_ascii=False);
        return result;
    
    def __str__(self):
        return self.to_json(self);


def process_xlsx(path):
    result = []
    wb = load_workbook(path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        xlsx_rows = []
        for row in ws.rows:
            lines = []
            for cell in row:
                lines.append(cell.value)
            xlsx_rows.append(lines)
        result.append(XslxVaue(sheet_name, xlsx_rows))
    return result


def value_to_file(value, destPath):
    print(value)
    print(destPath)
    f = open(destPath, 'w', encoding="utf-8")
    f.write(value)
    f.close()


def write_file(obj, destPath):
    result_json = json.dumps(obj, default=lambda o: o.__dict__, ensure_ascii=False, indent=2);
    value_to_file(result_json, destPath)

def xlsx_to_tx(xlsx_values):
    result = []
    for value in xlsx_values:
        course_name = value.sheet_name
        school_type_name = None
        txs = []
        for line in value.lines:
            ygtx, k1ktx = line
            ygtx = ygtx.strip() if ygtx != None else None
            k1ktx = k1ktx.strip() if k1ktx != None else None
            if ([ygtx, k1ktx] == [None, None]):
                if (len(txs) > 0):
                    result.append(TxTmp(school_type_name, course_name, txs))
                    txs = []
                continue
            if (ygtx != None and ygtx.find("备注") > -1):
                continue
            for tmp in school_type_names:
                if(ygtx and ygtx.find(tmp) > -1):
                    school_type_name = tmp
                    if (len(txs) > 0):
                        result.append(TxTmp(school_type_name, course_name, txs))
                        txs = []
                    break
            if (ygtx != None and (ygtx.find("央管") > -1 or ygtx.find("小学") > -1 or ygtx.find("初中") > -1 or ygtx.find("高中") > -1 or ygtx.startswith("     "))):
                continue
            if ygtx != None:
                ss = ygtx.split("、")
                for single_ygtx in ss:
                    txs.append(Tx(single_ygtx, k1ktx))
            else:
                txs.append(Tx(ygtx, k1ktx))
        if(len(txs) > 0):
            result.append(TxTmp(school_type_name, course_name, txs))
            txs = []
    return result
xlsxPath = 'F:/tmp/ygtk/考一考与央馆题型对应【除语文和英语】.xlsx'
jsonPath = 'F:/tmp/ygtk/txconvert.json'
result = process_xlsx(xlsxPath)
result_txs = xlsx_to_tx(result)
write_file(result_txs, jsonPath)
print('ok')
